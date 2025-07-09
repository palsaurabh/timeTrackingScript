'''
Task Logger Script

This Python script helps you track tasks throughout the day by creating or updating an Excel file for each day. The script performs the following operations:

1. Creates a new Excel file for the current day if it does not already exist.
2. Logs tasks with their start time, end time, and total time spent.
3. Prompts the user to log what they did since the last recorded task when the program is restarted.
4. Continuously prompts for new tasks, tracks their start time, and waits for the user to finish before recording the end time.

Dependencies:
- `openpyxl`: Used to manage Excel files. Install it using `pip install openpyxl`.

How to Use:
1. Run the script.
2. Follow the prompts to log your completed tasks and start new ones.
3. An Excel file named `tasks_YYYY-MM-DD.xlsx` will be created or updated in the current directory.

Author: Your Name
Date: YYYY-MM-DD
'''

import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
import time
import threading
import tkinter as tk
from tkinter import Label

# Function to calculate time difference in HH:MM:SS format
def calculate_duration(start, end):
    duration = end - start
    return str(duration)

# Function to create or load the daily Excel file
def get_excel_file():
    today = datetime.now().strftime('%Y-%m-%d')
    filename = f"tasks_{today}.xlsx"

    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "Tasks"
        ws.append(["Task Name", "Start Time", "End Time", "Total Time"])
        wb.save(filename)

    return filename

# Function to log a task into the Excel sheet
def log_task(filename, task_name, start_time, end_time):
    wb = load_workbook(filename)
    ws = wb.active

    total_time = calculate_duration(start_time, end_time)
    ws.append([task_name, start_time.strftime('%Y-%m-%d %H:%M:%S'), end_time.strftime('%Y-%m-%d %H:%M:%S'), total_time])
    print("Total Time: ", total_time)
    wb.save(filename)

# Function to get the last logged task's end time
def get_last_end_time(filename):
    wb = load_workbook(filename)
    ws = wb.active

    if ws.max_row > 1:
        last_end_time_str = ws.cell(row=ws.max_row, column=3).value
        return datetime.strptime(last_end_time_str, '%Y-%m-%d %H:%M:%S')
    
    return None

# Function to display elapsed time in the popup window
def update_time_popup(start_time, stop_event, label):
    while not stop_event.is_set():
        current_time = datetime.now()
        elapsed_seconds = (current_time - start_time).total_seconds()
        hours, remainder = divmod(elapsed_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)
        minutes_elapsed, k = divmod(elapsed_seconds, 60)
        print("Before try")
        try:
            # label.config(
            #     text=f"Elapsed time: {int(hours):02d} hours, {int(minutes):02d} minutes, {int(seconds):02d} seconds"
            # )
            label.config(
                text=f"{int(minutes_elapsed):02d}m"
            )
        except RuntimeError:
            #The window may gave been destroyed, exit the loop
            break
        
        print("Before sleep")
        time.sleep(1)
blink = 1      
def update_time_in_main_thread(start_time, stop_event, label, root, task_duration):
    if stop_event.is_set():
        return  # Stop updates if the event is set
    current_time = datetime.now()
    elapsed_seconds = (current_time - start_time).total_seconds()
    hours, remainder = divmod(elapsed_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    # label.config(
    #     text=f"Elapsed time: {int(hours):02d} hours, {int(minutes):02d} minutes, {int(seconds):02d} seconds"
    # )
    minutes_elapsed, k = divmod(elapsed_seconds, 60)
    label.config(
        text=f"{int(minutes_elapsed):02d}m"
    )

    global blink
    if minutes_elapsed > int(task_duration):
        if blink == 1:
            root.configure(bg="red")
            blink = 0
        else:
            root.configure(bg="green")
            blink = 1
    # Schedule the next update after 1 second
    root.after(1000, update_time_in_main_thread, start_time, stop_event, label, root, task_duration)

def create_time_popup(start_time, stop_event, task_name, task_duration):
    # Create a new Tkinter window
    root = tk.Tk()
    root.title(task_name)
    root.title("Timer")
    root.geometry("40x27")
    root.wm_attributes("-topmost", True)
    root.configure(bg="green")

    # Add a label to display elapsed time
    bold_font = ("Arial", 14, "bold")  # Bold font specification
    # label = Label(root, 
    #                 text="Elapsed time: 00:00:00", 
    #                 font=bold_font,
    #                 bg='red',
    #                 fg='black')
    label = Label(root, 
                    text="00m", 
                    font=bold_font,
                    bg='red',
                    fg='black')

    label.pack(expand=True)

    # Schedule the first update in the main thread
    root.after(1000, update_time_in_main_thread, start_time, stop_event, label, root, task_duration)

    def on_close():
        # Signal the thread to stop and destroy the Tkinter window
        stop_event.set()
        root.destroy()

    root.protocol("WM_DELETE_WINDOW", on_close)
    root.mainloop()

# Main script
def main():
    filename = get_excel_file()
    last_task_end_time = get_last_end_time(filename)

    if last_task_end_time:
        print("\n--- Logging Previous Task ---")
        task_done = input("What did you do since the last task? ")
        current_time = datetime.now()
        log_task(filename, task_done, last_task_end_time, current_time)
        print(f"Task '{task_done}' logged successfully.\n")

    while True:
        print("\n--- New Task ---")
        task_name = input("What are you going to do now? ")
        task_duration = input("Duration? ")
        start_time = datetime.now()
        print(f"Started task '{task_name}' at {start_time.strftime('%Y-%m-%d %H:%M:%S')}.")
        print("Press Enter when you finish the task...")

        # Create a stop event for the thread
        stop_event = threading.Event()

        popup_thread = threading.Thread(target=create_time_popup, args=(start_time, stop_event, task_name, task_duration))
        popup_thread.start()

        # Wait for the user to press Enter
        input()  # Blocks until Enter is pressed

        # Stop the popup thread
        if not stop_event.is_set():
            stop_event.set()
        popup_thread.join()

        # Record the end time
        end_time = datetime.now()

        # Calculate final elapsed time
        elapsed_seconds = (end_time - start_time).total_seconds()
        hours, remainder = divmod(elapsed_seconds, 3600)
        minutes, seconds = divmod(remainder, 60)

        # Log the task
        log_task(filename, task_name, start_time, end_time)
        print(f"\nTask '{task_name}' logged successfully.")
        print(
            f"Task '{task_name}' took {int(hours)} hours, {int(minutes)} minutes, and {int(seconds)} seconds.\n"
        )

        last_task_end_time = end_time

if __name__ == "__main__":
    main()
